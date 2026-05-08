import openpyxl
import sys

def move_data(src_sheet, dest_sheet, n):
    dest_row = 1
    for i in range(4 * n + 1, src_sheet.max_row + 1, 4):
        dest_sheet.cell(row=dest_row, column=1, value=src_sheet.cell(row=i, column=1).value)
        dest_sheet.cell(row=dest_row, column=2, value=src_sheet.cell(row=i, column=2).value)
        dest_sheet.cell(row=dest_row, column=3, value=src_sheet.cell(row=i+2, column=2).value)
        dest_sheet.cell(row=dest_row, column=4, value=src_sheet.cell(row=i+1, column=2).value)
        dest_sheet.cell(row=dest_row, column=5, value=src_sheet.cell(row=i, column=3).value)
        dest_row += 1

if len(sys.argv) < 3:
    print("用法: python excel_process.py <源Excel路径> <输出Excel路径> [起始偏移n]")
    raise SystemExit(1)

src_file_path = sys.argv[1]
dest_file_path = sys.argv[2]
n = int(sys.argv[3]) if len(sys.argv) > 3 else 0

src_workbook = openpyxl.load_workbook(src_file_path)
src_sheet = src_workbook.active

# 创建新的Excel文件
dest_workbook = openpyxl.Workbook()
dest_sheet = dest_workbook.active

# 调用函数进行数据移动
move_data(src_sheet, dest_sheet, n)

# 保存新的Excel文件
dest_workbook.save(dest_file_path)

print("数据移动完成并存入新的Excel文件。")
